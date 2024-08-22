from openpyxl.styles import Font, Border, Side, PatternFill, Alignment 
from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = 'Project'


# ========== 1줄씩 data 넣기
ws.append(["학번", "출석", "퀴즈1","퀴즈2", "중간고사", "기말고사", "프로젝트", "총점(SUM)", "성적"])


#10개 데이터 넣기 
for i in range(1,11): 
    ws.append([i,randint(1,10), randint(1,10), randint(1,10), randint(1,20), randint(1,30), randint(1,20)])
 
 
# 퀴즈2열만 10으로 데이터 다시채우고 
for row in ws.iter_rows(min_row=2):
    row[3].value = 10

sum = 0
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=9): # row 2~11줄까지만 
    for cell in row:
        
        # 각줄마다 sum 먼저 합산 부터 해주고 
        if isinstance(cell.value, int):
            sum += cell.value 
        
        # sum 을 총점 자리에 표시하고 
        jul = cell.row
        ws.cell(row = jul, column=8).value  = sum
        
        # 총점 점수별 = 성적 데이터 넣기 
        if sum  > 90 :
            ws.cell(row = jul, column=9).value  = 'A'
        elif sum  > 80 :
            ws.cell(row = jul, column=9).value  = 'B'
        elif sum  > 70 :
            ws.cell(row = jul, column=9).value  = 'C'
        else :
            ws.cell(row = jul, column=9).value  = 'D'
        
        # 근데 출석이 5미만 = 성적 무조건 'F'처리 까지 
        if ws.cell(row = jul, column=2).value < 5:
            
            # 출석 5도 레드 하고
            ws.cell(row = jul, column=2).font = Font(color="FF0000", italic=True, bold=True) # RGB 기준 빨강, 기울림, 굵게 
            
            # 성적에도 F처리, 배경색, 빨강
            ws.cell(row = jul, column=9).value  = 'F'
            ws.cell(row = jul, column=9).fill = PatternFill(fgColor="D4D4D4", fill_type="solid") # 배경을 그레이
            ws.cell(row = jul, column=9).font = Font(color="FF0000", italic=True, bold=True) # RGB 기준 빨강, 기울림, 굵게 
    print(sum)
    sum=0




# # ㅡㅡㅡㅡ 번호 칸 - 너비 좁혀보자!

# H열 너비 5로 설정
ws.column_dimensions["I"].width = 20

# 1행 높이를 50 으로 설정 
ws.row_dimensions[2].height = 50

for i in range(3,12):
    ws.row_dimensions[i].height = 30

# # ㅡㅡㅡㅡ 글자 스타일 적용


a1 = ws["A1"]
h1 = ws["H1"]
i1 = ws["I1"]

a1.font = Font(color="FF0000", italic=True, bold=True) # RGB 기준 빨강, 기울림, 굵게
h1.font = Font(color="CC33FF", name='Arial') # RGB 기준 보라, 폰트체
i1.font = Font(color = '0000FF', size=20, underline='single') # RGB 기준 파랑, 크게, 밑줄적용



# # ㅡㅡㅡㅡ 테두리 적용 
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
h1.border = thin_border
i1.border = thin_border



# ㅡㅡㅡㅡ 배경색 적용 / 정렬

# 90점 넘는 셀 - 초록색 

for row in ws.rows:
    for cell in row :
        # 각 cell 정렬 - center, left, right, top, bottom
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        
        if cell.column == 1:  
            cell.fill = PatternFill(fgColor="D4D4D4", fill_type="solid") # 배경을 그레이
        
        if cell.row == 1:  
            cell.fill = PatternFill(fgColor="E4F7BA", fill_type="solid") # 배경을 그레이
        
        

for row in ws.iter_rows( min_col=2, max_col=7):
    for cell in row :
        # cell 이 정수형 데이터이고 15점보다 높으면
        if isinstance(cell.value, int) and cell.value > 15 :
            cell.font = Font(color="0100FF") #폰트 색상 블루
            


ws.insert_rows(1)
ws.insert_cols(1) 


b15 = ws["B15"]
b15.font = Font(color = '0000FF', underline='single') # RGB 기준 파랑,
b15.value = ' ** 모든 cell 값 중 15이상은 파란색으로 표기했습니다.'

b17 = ws["B17"]
b17.font = Font(color = 'FF0000', underline='single') # RGB 기준 빨강,
b17.value = ' - 출석이 5미만인자는 강제 F 처리. -'


wb.save("rpa_project/1_excel/score.xlsx")