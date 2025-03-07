from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 데이터 넣기 
ws.append(('학번', '출석', '퀴즈1', '퀴즈2', '중간고사', '기말고사', '프로젝트'))

scores = [
    (1,10,8,5,14,26,12),
    (2,7,3,7,15,24,18),
    (3,9,5,8,8,12,4),
    (4,7,8,7,17,21,18),
    (5,7,8,7,16,25,15),
    (6,3,5,8,8,17,0),
    (7,4,9,10,16,27,18),
    (8,6,6,6,15,19,17),
    (9,10,10,9,19,30,19),
    (10,9,8,8,20,25,20)
    
]

for s in scores:
    ws.append(s)
    
    
# 퀴즈2 => 10으로 수정     
for idx, cell in enumerate(ws["D"]):
    if idx==0: # 제목은 skp
        continue
    cell.value = 10

ws["H1"] = "총점"
ws["I1"] = "성적"
    
for idx, score in enumerate(scores, start=2):
    sum_val = sum(score[1:]) - score[3] + 10 # 총점
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)
    # SUM(B2:G2)
    # SUM(B3:G3) ...
    
    
    # 총점별 성적
    grade = None
    if sum_val >= 90:
        grade = 'A'
    elif sum_val >= 80:
        grade = 'B'
    elif sum_val >= 70:
        grade = 'C'
    else:
        grade = 'D'
        
    # 출석 5점 미만 = 성적 무조건 F
    if score[1] < 5:
        grade = "F"
        
        
    ws.cell(row=idx, column=9).value = grade # I열에 성적정보 
    
    
    
wb.save("rpa_project/1_excel/quiz.xlsx")