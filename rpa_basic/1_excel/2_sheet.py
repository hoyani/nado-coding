from openpyxl import Workbook
wb = Workbook()

#wb.active()
ws = wb.create_sheet() # 새로운 sheet 기본이름으로 생성 => sheet1로 추가됌!! 
ws.title = "MySheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "d2ecf9" # 탭 색상 변경 (RGB 형태 값) - w3school

# ㅡㅡㅡㅡ Sheet, Mysheet, Yoursheet
ws1 = wb.create_sheet("Yoursheet") # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index에 sheet 생성


new_ws = wb["NewSheet"] # Dict 형태로 sheet 에 접근 

print(wb.sheetnames) # 모든 sheet 이름 확인 

# ㅡㅡㅡㅡ sheet 복사
new_ws["A1"] = "Test" #A1셀에다 값 입력
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"


wb.save("rpa_basic/1_excel/sample.xlsx")

