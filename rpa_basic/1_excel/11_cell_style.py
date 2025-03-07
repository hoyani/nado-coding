from openpyxl.styles import Font, Border, Side, PatternFill, Alignment 
from openpyxl import load_workbook
wb = load_workbook("rpa_basic/1_excel/sample.xlsx")

ws = wb.active

# 번호, 영어, 수학
a1 = ws["A1"] # 번호
b1 = ws["B1"] # 영어
c1 = ws["C1"] # 수학


# ㅡㅡㅡㅡ 번호 칸 - 너비 좁혀보자!

# A열 너비 5로 설정
ws.column_dimensions["A"].width = 5 

# 1행 높이를 50 으로 설정 
ws.row_dimensions[1].height = 50

# ㅡㅡㅡㅡ 글자 스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True) # RGB 기준 빨강, 기울림, 굵게
b1.font = Font(color="CC33FF", name='Arial', strike=True) # RGB 기준 보라, 폰트체, 취소선
c1.font = Font(color = '0000FF', size=20, underline='single') # RGB 기준 파랑, 크게, 밑줄적용


# ㅡㅡㅡㅡ 테두리 적용 
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border



# ㅡㅡㅡㅡ 배경색 적용 / 정렬

# 90점 넘는 셀 - 초록색 

for row in ws.rows:
    for cell in row :
        
        # 각 cell 정렬 - center, left, right, top, bottom
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        
        if cell.column == 1: # A 번호열은 제외 
            continue
        
        # cell 이 정수형 데이터이고 90점보다 높으면
        if isinstance(cell.value, int) and cell.value > 90 :
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") # 배경을 초록색으로
            cell.font = Font(color="FF0000") #폰트 색상 변경 
            
        
# ㅡㅡㅡㅡ 틀 고정
ws.freeze_panes = "B2" # B2 기준으로 틀 고정 - 아래스크롤이든, 오른쪽스크롤이든 


wb.save("rpa_basic/1_excel/sample_style.xlsx") 