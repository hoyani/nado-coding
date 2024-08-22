from openpyxl import load_workbook


# < print 터미널에 표시해보기 >


# ㅡㅡㅡㅡ 수식 그대로 가져오는 상태 

wb= load_workbook("rpa_basic/1_excel/sample_formula.xlsx")
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell)

# ㅡㅡㅡㅡ 수식 계산된 상태

# ㅡ1) 숫자는 되는데 
# ㅡ2) 수식 부분은 ,,,

# evaluate 게산되지않은 상태의 데이터 : 터미널엔 none 이라 표시 
# 수식이 올라간 엑셀파일 닫을때 저장하면 = 엑셀에 게산된 값으로 남음 / 이제 다시 실행하면 : 터미널에 그 남은 게산된 값이 보임! 
# => 즉 열때 '최종 산출물'을 여는용으로 이 파이썬 쓰는게 좋다...!!

wb= load_workbook("rpa_basic/1_excel/sample_formula.xlsx", data_only=True) # <-- 여기 속성 하나 추가하니!!  이 차이 
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell)