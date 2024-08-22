from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells("B2:D2") # B2 부터 D2 까지 합치겟다
ws["B2"].value = "Merged Cell"

wb.save("rpa_basic/1_excel/sample_merge.xlsx")