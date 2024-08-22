from openpyxl import load_workbook
wb = load_workbook("rpa_basic/1_excel/sample.xlsx")

ws = wb.active


# ㅡㅡㅡㅡ 이동하자 

# 번호 영어 수학
# 번호 (국어) 영어 수학

ws.move_range("B1:C11", rows=0, cols=1) 
# 대상 : b1:c11 얘네를...
# 목적지 : row,cols 정보 <- 여기로 이동할거다!! / 즉, 그대로 한 열만 옆으로 이동하겟단거 ㅋㅋ 
# ㄴ 이동 됏고..

# ㅡㅡㅡㅡ 이제 빈자리에 추가 
ws['B1'].value = '국어' # 새 셀 추가 




# ㅡㅡㅡㅡ 이동하자 2

# 번호 영어 수학 

ws.move_range("C1:C11", rows=5, cols=-1)  #역이니 덮어쓰기 되겟지 ㅋㅋ




wb.save('rpa_basic/1_excel/sample_korean.xlsx')