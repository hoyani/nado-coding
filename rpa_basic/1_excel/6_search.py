from openpyxl import load_workbook
wb = load_workbook("rpa_basic/1_excel/sample.xlsx") #기존 파일 불러오기

ws = wb.active


for row in ws.iter_rows(min_row=2):
    #번호, 영어, 수학
    if int(row[1].value) > 80:
        print(row[0].value, '번 학생은 영어천재')
        

# 문제상황? 영어 말고 컴터엿다, 변경해야!
for row in ws.iter_rows(max_row=1): #제목부터 ~
    for cell in row:
        if cell.value == '영어' : 
            cell.value = "컴퓨터"
            

wb.save("rpa_basic/1_excel/sample_modified.xlsx")