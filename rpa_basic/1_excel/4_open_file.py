from openpyxl import load_workbook # 파일 불러오기 
wb = load_workbook("rpa_basic/1_excel/sample.xlsx") # sample.xlsx 파일에서 wb을 불러옴
ws = wb.active # 활성화된 sheet 


# ㅡㅡㅡㅡ cell 데이터 불러오기 
# for x in range(1,11):
#     for y in range(1,11):
#         print(ws.cell(row=x, column=y).value, end=" ") # 1,2,3,4 ...
#     print()
    
# ㅡㅡㅡㅡ cell 갯수를 모를 때 
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column +1):
        print(ws.cell(row=x, column=y).value, end=" ") # 1,2,3,4 ...
    print()
    