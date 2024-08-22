from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# ============================== 1줄씩 data 넣기
ws.append(["번호", "영어", "수학"])

for i in range(1,11): #10개 데이터 넣기 
    ws.append([i,randint(0,100), randint(0,100)])
    

# ============================== 값 셀렉하기 

# ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ 컬럼(열) 단위

# ㅡㅡㅡㅡ 1) 열 1개

col_B = ws["B"] # 영어 column 만 가지고 오기 
#print(col_B) 
for cell in col_B :
    print(cell.value)
    
'''
결과
ㄴ
영어
3
68
77
6
82
33
39
68
70
29

'''


# ㅡㅡ 2) 열 여러개
    
    
col_range = ws["B:C"] # 영어~수학 컬럼까지 함꼐 갖고 오기 
for cols in col_range :
    for cell in cols :
        print(cell.value)
        
'''
결과
ㄴ
영어
3
68
77
6
82
33
39
68
70
29
수학
56
59
60
0
3
51
1
42
62
68
'''
        
# ㅡㅡ 방법1 : tuple 

# 모든 셀 출력 -> 셀 정보값만 o 값은 x 출력
print(tuple(ws.columns))
'''
결과 : 
(<Cell 'Sheet'.A1>, ~ A11 까지
(<Cell 'Sheet'.B1>, ~ B11 까지
(<Cell 'Sheet'.C1>, ~ C11 까지

'''

# 모든 열중에서 -> row index 1번(둘째줄) 만 값 출력

for column in tuple(ws.columns):
    print(column[1].value, end='ㅡ') 
print()

'''
결과
1ㅡ42ㅡ2ㅡ (모든열, 첫째줄)
'''



# ㅡㅡ 방법2 : iter_col 함수


# 모든 열중에서 -> row index 1번(둘째줄) 만 값 출력

for column in ws.iter_cols(): #전체 column
    print(column[1].value)
'''
결과
ㄴ

1
42
2 (모든열, 첫째줄)
'''




# 특정 셀 범위 지정 -> 특정 row (cell) 값 출력  

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3): # row 1~5줄까지만 / col 1~3열 (수학,영어 여기서지정 )까지 slice
    print(col[0].value, col[1].value) # 이미 정제된 새 틀안에서의 index다!
    #print(col)

'''
결과
ㄴ
번호 1
영어 42
수학 2
'''

# ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ 줄 단위
    
    
# ㅡㅡ 1) 줄 1개
row_title = ws[1] # 1번째 row 만 가지고 오기 
for cell in row_title:
    print(cell.value, end=' ')  
print()

'''
결과
ㄴ
번호 영어 수학 

'''

# ㅡㅡ 2) 줄 여러개 (특정~특정)

row_range = ws[2:6] # 2번째줄 ~ 6번째 줄까지 가지고오기 (일반 슬라이싱이랑 조금 다름) / title 1번줄 제외 출력
for rows in row_range:
    for cell in rows :
        print(cell.value, end=' ')
    print()
'''
결과
1 42 2 
2 15 37 
3 46 60 
4 87 2 
5 38 69 
'''



# ㅡㅡ 2) 줄 여러개 (특정~끝까지)

from openpyxl.utils.cell import coordinate_from_string


row_range = ws[2:ws.max_row] #2번줄부터 ~ 마지막 줄까지
for rows in row_range:
    for cell in rows:
        # print(cell.value, end=' ')
        # print(cell.coordinate, end=' ') #A10 등 위치값 
        
        xy = coordinate_from_string(cell.coordinate) #숫자 문자 분리 
        # print(xy, end=' ')
        print(xy[0], end='')
        print(xy[1], end=' ')    
    print()

'''
ㅡㅡ결과ㅡㅡㄱ
A2 B2 C2 |
A3 B3 C3 |
A4 B4 C4 |
A5 B5 C5 
A6 B6 C6 
A7 B7 C7 
A8 B8 C8 
A9 B9 C9 
A10 B10 C10 
A11 B11 C11 
'''

# ㅡㅡ 3) 전체 Rows


# cell 정보값만 출력 
print(tuple(ws.rows))
'''
결과
ㄴ
(<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>), 
(<Cell 'Sheet'.A2>, <Cell 'Sheet'.B2>, <Cell 'Sheet'.C2>), 
..
(<Cell 'Sheet'.A11>, <Cell 'Sheet'.B11>, <Cell 'Sheet'.C11>), 11 까지

'''



# ㅡㅡ 방법1) tuple 이용


# 모든 row 의 -> 특정 column의 값을 출력 

for row in tuple(ws.rows):
    print(row[2].value)
'''
결과
ㄴ
수학
3
67
9
69
37
80
93
66
29
3
'''

# ㅡㅡ 방법2) iter함수 이용


# 모든 row 의 -> 특정 column의 값을 출력 

for row in ws.iter_rows(): #전체 row
    print(row[1].value)  
'''
결과
ㄴ
영어
59
20
33
90
41
46
67
59
13
69
'''


# 특정 셀 범위 지정 -> 특정 row (cell) 값 출력  

for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3): # row 2~11줄까지만 / col 2~3열 (수학,영어 여기서지정 )까지 slice
    print(row[0].value, row[1].value) # 이미 정제된 새 틀안에서의 index다!
    #print(row) # cell 정보값 출력 

'''
결과
ㄴ
96 64
73 88
20 68
15 64
12 69
83 20
71 79
42 79
32 87
28 44
'''

wb.save("rpa_basic/1_excel/sample.xlsx")